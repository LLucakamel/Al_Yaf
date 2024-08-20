from django.shortcuts import render, redirect, get_object_or_404  # Added get_object_or_404 here
from .models import Project
from .forms import ProjectForm, ProjectPhaseForm, EditPhaseForm  # Import EditPhaseForm
from django.http import JsonResponse, HttpResponse  # Add HttpResponse to the existing import
from openpyxl import Workbook, load_workbook  # Add load_workbook to the existing import
from io import BytesIO  # Import BytesIO for generating Excel file

def project_create(request):
    if request.method == "POST":
        name = request.POST.get('name')
        code = request.POST.get('code')
        location = request.POST.get('location')
        consultant = request.POST.get('consultant')
        Project.objects.create(name=name, code=code, location=location, consultant=consultant)
        return redirect('Projects:project-list')
    return render(request, 'projects.html')

def project_list(request):
    projects = Project.objects.all()
    return render(request, 'project_list.html', {'projects': projects})

def add_project(request):
    if request.method == 'POST':
        form = ProjectForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('Projects:project-list')  # Redirect to the project list after saving
    else:
        form = ProjectForm()
    return render(request, 'add_project.html', {'form': form})

def product_details(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    return JsonResponse({
        'name': project.name,
        'code': project.code,
        'location': project.location,
        'consultant': project.consultant
    })

def edit_project(request, project_id):
    project = get_object_or_404(Project, pk=project_id)
    if request.method == 'POST':
        form = ProjectForm(request.POST, instance=project)
        if form.is_valid():
            form.save()
            return redirect('Projects:project-list')
    else:
        form = ProjectForm(instance=project)
    return render(request, 'edit_project.html', {'form': form})

def import_projects(request):
    if request.method == 'POST' and request.FILES['project_file']:
        excel_file = request.FILES['project_file']
        wb = load_workbook(filename=excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
            Project.objects.update_or_create(
                name=row[0],
                defaults={
                    'code': row[1],
                    'location': row[2],
                    'consultant': row[3]
                }
            )

        return redirect('Projects:project-list')  # Redirect to the project list after importing

    return render(request, 'import_projects.html')

def export_projects(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=projects.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # Column headers
    columns = ['Name', 'Code', 'Location', 'Consultant']
    ws.append(columns)

    # Data rows
    projects = Project.objects.all()
    for project in projects:
        ws.append([project.name, project.code, project.location, project.consultant])

    wb.save(response)
    return response

def search_projects(request):
    query = request.POST.get('q', '')
    projects = Project.objects.filter(name__icontains=query)  # Adjust the filter as needed
    return render(request, 'projects/project_list.html', {'projects': projects})

# New code for adding project phases

def add_project_phase(request):
    if request.method == 'POST':
        form = ProjectPhaseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('Projects:project_phase_list')  # Redirect to the list of project phases
    else:
        form = ProjectPhaseForm()
    return render(request, 'add_project_phase.html', {'form': form})

from django.views.generic import ListView
from .models import ProjectPhase

class ProjectPhaseListView(ListView):
    model = ProjectPhase
    template_name = 'project_phase_list.html'
    context_object_name = 'phases'

def project_phase_list_api(request):
    phases = ProjectPhase.objects.all().values('id', 'phase_name')
    return JsonResponse(list(phases), safe=False)

def edit_phase(request, id):
    phase = get_object_or_404(ProjectPhase, id=id)
    if request.method == 'POST':
        form = EditPhaseForm(request.POST, instance=phase)
        if form.is_valid():
            form.save()
            return redirect('Projects:project_phase_list')  # Redirect to the phase list
    else:
        form = EditPhaseForm(instance=phase)
    return render(request, 'edit_phase.html', {'form': form})

def remove_phase(request, phase_id):
    # Logic to delete a project phase
    phase = get_object_or_404(ProjectPhase, id=phase_id)
    phase.delete()
    return redirect('Projects:project_phase_list')